import os
import pandas as pd
from openpyxl import Workbook
import re
import openpyxl

def save_employee_data(employee_data, employee_name, month_year, salary_table, output_folder):
    # Tạo tên file xuất
    sanitized_employee_name = re.sub(r'[\\/:"*?<>|]+', "_", employee_name)
    file_name = f'{sanitized_employee_name}_{month_year}_{salary_table}.xlsx'
    output_path = os.path.join(output_folder, file_name)

    # Tạo workbook mới và viết dữ liệu vào file
    wb = Workbook()
    ws = wb.active
    ws.title = "Thông Tin Nhân Viên"
    
    # Ghi tiêu đề cột
    ws.append(list(employee_data.columns))
    
    # Ghi dữ liệu nhân viên
    for row in employee_data.itertuples(index=False, name=None):
        ws.append(row)
    
    # Điều chỉnh chiều rộng cột
    min_width = 15
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max(max_length + 2, min_width)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Lưu file Excel
    wb.save(output_path)
    print(f"Thông tin nhân viên {employee_name} đã được lưu tại: {output_path}")

def load_data(file_path, sheet_name):
    # Đọc file excel và lấy dữ liệu của sheet
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    
    # Lấy thông tin tháng và năm từ các cột F, G, H, I tại dòng 1
    month_info = df.iloc[0, 5:9].values  # Cột F, G, H, I
    month_year = ' '.join([str(i) for i in month_info if pd.notna(i)]).strip()
    print(f"Tháng và năm: {month_year}")
    
    # In toàn bộ dữ liệu dòng 8 (header của bảng lương)
    header_row = df.iloc[7].dropna().values  # Lấy tất cả các giá trị không phải NaN
    header_row = [str(i).strip() for i in header_row if str(i).strip()]  # Loại bỏ khoảng trống
    print(f"Dữ liệu dòng 8 (header): {header_row}")
    
    # Tìm tất cả các bảng lương có chứa chuỗi 'BẢNG LƯƠNG'
    salary_tables = [cell for cell in header_row if 'BẢNG LƯƠNG' in cell]
    
    if not salary_tables:
        print("Không tìm thấy bảng lương nào theo cú pháp 'BẢNG LƯƠNG'.")
    else:
        print(f"Các bảng lương tìm thấy: {salary_tables}")
    
    # Đặt tên cột từ dòng 9
    df.columns = df.iloc[8]
    # Loại bỏ các dòng đầu tiên trước dữ liệu thực
    df = df.drop(index=range(9))

    return df, month_year, salary_tables

def clean_data(df):
    # Loại bỏ các cột toàn NaN
    df = df.dropna(how='all', axis=1)

    # Giữ lại tất cả các hàng, chỉ xóa NaN trong từng hàng
    df = df.fillna('')  # Thay thế NaN bằng chuỗi rỗng

    # In ra dữ liệu đã làm sạch
    print("Dữ liệu sau khi làm sạch:")
    print(df.head(10))  # In ra 10 dòng đầu tiên để kiểm tra

    return df

def search_employee(data, keyword, salary_table, table_start_col, table_end_col):
    # Lọc cột thuộc bảng lương được chọn (từ table_start_col đến table_end_col)
    # Giữ lại cột A và B (cột Họ tên NV và các thông tin chung)
    relevant_data = pd.concat([data.iloc[:, 0:19], data.iloc[:, table_start_col:table_end_col]], axis=1)
    
    name_column = 'Họ tên NV'
    
    try:
        # Đảm bảo cột 'Họ tên NV' là kiểu chuỗi
        relevant_data[name_column] = relevant_data[name_column].astype(str)
        
        # Tìm kiếm nhân viên theo tên hoặc ID
        result = relevant_data[relevant_data[name_column].str.contains(keyword, case=False, na=False)]
        
        # Kiểm tra nếu tìm thấy nhân viên
        if not result.empty:
            print(f"Thông tin nhân viên tìm thấy trong {salary_table}:")
            print(result)
        else:
            print(f"Không tìm thấy nhân viên trong {salary_table}.")
        
    except KeyError as e:
        print(f"Lỗi: Không tìm thấy cột '{name_column}' trong dữ liệu.")
        return None
    
    return result

    # Tạo tên file xuất
    sanitized_employee_name = re.sub(r'[\\/:"*?<>|]+', "_", employee_name)
    file_name = f'{sanitized_employee_name}_{month_year}_{salary_table}.xlsx'
    output_path = os.path.join(output_folder, file_name)

    # Tạo workbook mới và viết dữ liệu vào file
    wb = Workbook()
    ws = wb.active
    ws.title = "Thông Tin Nhân Viên"
    
    # Ghi tiêu đề cột
    ws.append(list(employee_data.columns))
    
    # Ghi dữ liệu nhân viên
    for row in employee_data.itertuples(index=False, name=None):
        ws.append(row)
    
    # Định dạng tiêu đề
    for cell in ws[1]:
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Sao chép màu sắc và định dạng từ sheet gốc cho từng ô
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(employee_data)+1), start=2):
        for col_idx, cell in enumerate(row, start=1):
            original_cell = original_sheet.cell(row=row_idx, column=col_idx)
            
            # Sao chép font, fill và alignment từ ô gốc
            cell.font = Font(
                name=original_cell.font.name,
                size=original_cell.font.size,
                bold=original_cell.font.bold,
                italic=original_cell.font.italic,
                vertAlign=original_cell.font.vertAlign,
                underline=original_cell.font.underline,
                strike=original_cell.font.strike,
                color=original_cell.font.color
            )
            cell.fill = PatternFill(
                fill_type=original_cell.fill.fill_type,
                start_color=original_cell.fill.start_color.rgb if original_cell.fill.start_color else None,
                end_color=original_cell.fill.end_color.rgb if original_cell.fill.end_color else None
            )
            cell.alignment = Alignment(
                horizontal=original_cell.alignment.horizontal,
                vertical=original_cell.alignment.vertical
            )
            
            # Định dạng các ô là số
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
    
    # Điều chỉnh chiều rộng cột
    min_width = 15
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max(max_length + 2, min_width)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Lưu file Excel
    wb.save(output_path)
    print(f"Thông tin nhân viên {employee_name} đã được lưu tại: {output_path}")
def main():
    excel_file = r'D:\Salary_Data\2024-_09_BangLuong_v01.xlsx'
    base_folder = r'D:\Salary_Data'

    # Đọc danh sách các sheet trong file Excel
    xls = pd.ExcelFile(excel_file)
    sheets = xls.sheet_names
    print("Danh sách bảng lương có sẵn:")
    for i, sheet in enumerate(sheets, 1):
        print(f"{i}. {sheet}")

    sheet_input = input("Chọn số thứ tự bảng lương: ").strip()
    if sheet_input.isdigit() and 1 <= int(sheet_input) <= len(sheets):
        sheet_name = sheets[int(sheet_input) - 1]
    else:
        print("Vui lòng chọn số hợp lệ.")
        return

    try:
        # Load dữ liệu và lấy thông tin tháng/năm và các bảng lương
        data, month_year, salary_tables = load_data(excel_file, sheet_name)
        cleaned_data = clean_data(data)

        if not salary_tables:
            print("Không tìm thấy bảng lương nào theo cú pháp 'BẢNG LƯƠNG'.")
            return
        
        print("Các bảng lương có sẵn:")
        for i, table in enumerate(salary_tables, 1):
            print(f"{i}. {table}")
        
        # Nhập tên hoặc ID nhân viên
        keyword = input("Nhập tên hoặc ID nhân viên: ").strip()
        
        for table in salary_tables:
            if table == 'BẢNG LƯƠNG KỲ 01':
                table_start_col = 19  # Cột T (19)
                table_end_col = 55  # Cột BC (55)
            elif table == 'BẢNG LƯƠNG KỲ 02':
                table_start_col = 56  # Cột BD (56)
                table_end_col = 94  # Cột CR (94)
            
            # Tìm kiếm trong từng bảng lương
            employee_data = search_employee(cleaned_data, keyword, table, table_start_col, table_end_col)
            if employee_data is not None and not employee_data.empty:
                employee_name = employee_data.iloc[0]['Họ tên NV']
                output_folder = os.path.join(base_folder, f'{month_year}_{table}')
                if not os.path.exists(output_folder):
                    os.makedirs(output_folder)

                # Lưu dữ liệu nhân viên (không copy màu sắc hay định dạng)
                save_employee_data(employee_data, employee_name, month_year, table, output_folder)
    except Exception as e:
        print(f"Lỗi: {e}")

if __name__ == "__main__":
    main()